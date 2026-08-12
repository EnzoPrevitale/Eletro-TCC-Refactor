import tkinter as tk
from tkinter import ttk, messagebox
import serial
import serial.tools.list_ports
import threading
import time
import math
import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg


# ============================================================
# CONFIGURAÇÕES
# ============================================================

PORTA = "COM5"
BAUDRATE = 115200

PASTA_DADOS = r"C:\Users\jayan\OneDrive\Área de Trabalho\TCC_RC\dados"

VCC = 5.0

TENSAO_MAX = 4.99
TENSAO_MIN = 0.005

MAX_TABELA = 1000


# ============================================================
# VARIÁVEIS
# ============================================================

arduino = None

teste_ativo = False
thread_leitura = None

dados_excel = []

dados_carga = []
dados_descarga = []

ciclo_atual = 0
estado_atual = "AGUARDANDO"

ultimo_ciclo_grafico = 0

tau = 0
tempo_5tau = 0
tempo_ciclo = 0
tempo_estimado = 0


# ============================================================
# JANELA
# ============================================================

janela = tk.Tk()

janela.title(
    "Sistema de Teste de Circuito RC"
)

janela.geometry(
    "1400x850"
)

janela.minsize(
    1100,
    700
)


# ============================================================
# ESTILO
# ============================================================

style = ttk.Style()

try:
    style.theme_use("clam")
except:
    pass

style.configure(
    "TNotebook.Tab",
    font=("Arial", 11, "bold"),
    padding=[15, 8]
)

style.configure(
    "Titulo.TLabel",
    font=("Arial", 18, "bold")
)


# ============================================================
# NOTEBOOK
# ============================================================

notebook = ttk.Notebook(janela)

notebook.pack(
    fill="both",
    expand=True,
    padx=10,
    pady=10
)


# ============================================================
# ABA CONFIGURAÇÃO
# ============================================================

aba_config = ttk.Frame(notebook)

notebook.add(
    aba_config,
    text="Configuração"
)


titulo = ttk.Label(
    aba_config,
    text="CONFIGURAÇÃO DO ENSAIO RC",
    style="Titulo.TLabel"
)

titulo.pack(pady=20)


# ============================================================
# COMPONENTES
# ============================================================

frame_componentes = ttk.LabelFrame(
    aba_config,
    text="Parâmetros do circuito RC",
    padding=20
)

frame_componentes.pack(
    fill="x",
    padx=30,
    pady=10
)


ttk.Label(
    frame_componentes,
    text="Capacitor (µF):"
).grid(
    row=0,
    column=0,
    padx=10,
    pady=10
)


entrada_capacitor = ttk.Entry(
    frame_componentes,
    width=20
)

entrada_capacitor.insert(
    0,
    "470"
)

entrada_capacitor.grid(
    row=0,
    column=1,
    padx=10,
    pady=10
)


ttk.Label(
    frame_componentes,
    text="Resistor (Ω):"
).grid(
    row=1,
    column=0,
    padx=10,
    pady=10
)


entrada_resistor = ttk.Entry(
    frame_componentes,
    width=20
)

entrada_resistor.insert(
    0,
    "2200"
)

entrada_resistor.grid(
    row=1,
    column=1,
    padx=10,
    pady=10
)


# ============================================================
# CÁLCULOS
# ============================================================

frame_calculos = ttk.LabelFrame(
    aba_config,
    text="Cálculos teóricos",
    padding=20
)

frame_calculos.pack(
    fill="x",
    padx=30,
    pady=10
)


label_tau = ttk.Label(
    frame_calculos,
    text="τ teórico: -- s",
    font=("Arial", 11)
)

label_tau.grid(
    row=0,
    column=0,
    padx=10,
    pady=5
)


label_5tau_carga = ttk.Label(
    frame_calculos,
    text="5τ carga: -- s",
    font=("Arial", 11)
)

label_5tau_carga.grid(
    row=0,
    column=1,
    padx=30,
    pady=5
)


label_5tau_descarga = ttk.Label(
    frame_calculos,
    text="5τ descarga: -- s",
    font=("Arial", 11)
)

label_5tau_descarga.grid(
    row=0,
    column=2,
    padx=30,
    pady=5
)


label_ciclo = ttk.Label(
    frame_calculos,
    text="1 ciclo: -- s",
    font=("Arial", 11, "bold")
)

label_ciclo.grid(
    row=1,
    column=0,
    padx=10,
    pady=10
)


label_tempo_total = ttk.Label(
    frame_calculos,
    text="Tempo total: --",
    font=("Arial", 11, "bold")
)

label_tempo_total.grid(
    row=1,
    column=1,
    columnspan=2,
    padx=20,
    pady=10
)


# ============================================================
# MODO DO ENSAIO
# ============================================================

frame_modo = ttk.LabelFrame(
    aba_config,
    text="Planejamento do ensaio",
    padding=20
)

frame_modo.pack(
    fill="x",
    padx=30,
    pady=10
)


modo_var = tk.StringVar(
    value="ciclos"
)


ttk.Radiobutton(
    frame_modo,
    text="Quantidade de ciclos",
    variable=modo_var,
    value="ciclos",
    command=lambda: atualizar_calculos()
).grid(
    row=0,
    column=0,
    padx=10
)


ttk.Radiobutton(
    frame_modo,
    text="Tempo de ensaio",
    variable=modo_var,
    value="tempo",
    command=lambda: atualizar_calculos()
).grid(
    row=0,
    column=1,
    padx=10
)


ttk.Label(
    frame_modo,
    text="Quantidade:"
).grid(
    row=1,
    column=0,
    padx=10,
    pady=15
)


entrada_quantidade = ttk.Entry(
    frame_modo,
    width=15
)

entrada_quantidade.insert(
    0,
    "30"
)

entrada_quantidade.grid(
    row=1,
    column=1,
    padx=10,
    pady=15
)


ttk.Label(
    frame_modo,
    text="ciclos / minutos"
).grid(
    row=1,
    column=2,
    padx=10
)


label_tempo_estimado = ttk.Label(
    aba_config,
    text="Tempo estimado: --",
    font=("Arial", 13, "bold")
)

label_tempo_estimado.pack(
    pady=15
)


# ============================================================
# STATUS
# ============================================================

label_status = ttk.Label(
    aba_config,
    text="STATUS: MONITORANDO ARDUINO",
    font=("Arial", 13, "bold")
)

label_status.pack(
    pady=10
)


# ============================================================
# BOTÕES
# ============================================================

frame_botoes = ttk.Frame(
    aba_config
)

frame_botoes.pack(
    pady=20
)


def criar_botao(
    parent,
    texto,
    comando
):
    return tk.Button(
        parent,
        text=texto,
        command=comando,
        font=("Arial", 12, "bold"),
        width=16,
        height=2,
        relief="raised",
        bd=3,
        cursor="hand2"
    )


botao_parar = criar_botao(
    frame_botoes,
    "■ PARAR",
    lambda: parar_teste()
)

botao_parar.grid(
    row=0,
    column=0,
    padx=10
)


botao_novo = criar_botao(
    frame_botoes,
    "NOVO TESTE",
    lambda: novo_teste()
)

botao_novo.grid(
    row=0,
    column=1,
    padx=10
)


botao_salvar = criar_botao(
    frame_botoes,
    "SALVAR EXCEL",
    lambda: salvar_excel()
)

botao_salvar.grid(
    row=0,
    column=2,
    padx=10
)


# ============================================================
# ABA GRÁFICOS
# ============================================================

aba_graficos = ttk.Frame(
    notebook
)

notebook.add(
    aba_graficos,
    text="Gráficos"
)


figura = Figure(
    figsize=(10, 7),
    dpi=100
)


ax_carga = figura.add_subplot(211)

ax_descarga = figura.add_subplot(212)


canvas = FigureCanvasTkAgg(
    figura,
    master=aba_graficos
)

canvas.get_tk_widget().pack(
    fill="both",
    expand=True
)


# ============================================================
# ABA TABELA
# ============================================================

aba_dados = ttk.Frame(
    notebook
)

notebook.add(
    aba_dados,
    text="Tabela"
)


frame_tabela = ttk.Frame(
    aba_dados
)

frame_tabela.pack(
    fill="both",
    expand=True,
    padx=20,
    pady=20
)


colunas = (
    "Ciclo",
    "Estado",
    "Tempo (s)",
    "Tensão (V)"
)


tabela = ttk.Treeview(
    frame_tabela,
    columns=colunas,
    show="headings"
)


for coluna in colunas:

    tabela.heading(
        coluna,
        text=coluna
    )

    tabela.column(
        coluna,
        width=180,
        anchor="center"
    )


scrollbar = ttk.Scrollbar(
    frame_tabela,
    orient="vertical",
    command=tabela.yview
)


tabela.configure(
    yscrollcommand=scrollbar.set
)


tabela.pack(
    side="left",
    fill="both",
    expand=True
)


scrollbar.pack(
    side="right",
    fill="y"
)


# ============================================================
# BOTÃO IR PARA O FINAL
# ============================================================

def ir_para_final():

    itens = tabela.get_children()

    if itens:

        tabela.see(
            itens[-1]
        )


botao_final = tk.Button(
    aba_dados,
    text="▼ IR PARA O FINAL",
    command=ir_para_final,
    font=("Arial", 11, "bold"),
    width=25,
    height=2
)

botao_final.pack(
    pady=(0, 15)
)


# ============================================================
# CÁLCULOS
# ============================================================

def atualizar_calculos():

    global tau
    global tempo_5tau
    global tempo_ciclo
    global tempo_estimado

    try:

        C = float(
            entrada_capacitor
            .get()
            .replace(",", ".")
        )

        R = float(
            entrada_resistor
            .get()
            .replace(",", ".")
        )

        tau = (
            R * C
        ) / 1000000.0

        tempo_5tau = (
            5 * tau
        )

        tempo_ciclo = (
            tempo_5tau * 2
        )

        label_tau.config(
            text=f"τ teórico: {tau:.3f} s"
        )

        label_5tau_carga.config(
            text=f"5τ carga: {tempo_5tau:.3f} s"
        )

        label_5tau_descarga.config(
            text=f"5τ descarga: {tempo_5tau:.3f} s"
        )

        label_ciclo.config(
            text=f"1 ciclo: {tempo_ciclo:.3f} s"
        )

        if modo_var.get() == "ciclos":

            quantidade = int(
                entrada_quantidade.get()
            )

            tempo_estimado = (
                quantidade *
                tempo_ciclo
            )

            minutos = int(
                tempo_estimado // 60
            )

            segundos = int(
                tempo_estimado % 60
            )

            label_tempo_total.config(
                text=(
                    f"Tempo total: "
                    f"{minutos} min "
                    f"{segundos} s"
                )
            )

            label_tempo_estimado.config(
                text=(
                    f"{quantidade} ciclos → "
                    f"{minutos} min "
                    f"{segundos} s"
                )
            )

        else:

            minutos = float(
                entrada_quantidade
                .get()
                .replace(",", ".")
            )

            tempo_estimado = (
                minutos * 60
            )

            ciclos = int(
                tempo_estimado /
                tempo_ciclo
            )

            label_tempo_total.config(
                text=(
                    f"Tempo de ensaio: "
                    f"{minutos:g} min"
                )
            )

            label_tempo_estimado.config(
                text=(
                    f"{minutos:g} min → "
                    f"aprox. {ciclos} ciclos"
                )
            )

    except:

        label_tau.config(
            text="τ teórico: --"
        )

        label_5tau_carga.config(
            text="5τ carga: --"
        )

        label_5tau_descarga.config(
            text="5τ descarga: --"
        )

        label_ciclo.config(
            text="1 ciclo: --"
        )

        label_tempo_total.config(
            text="Tempo total: --"
        )

        label_tempo_estimado.config(
            text="Tempo estimado: --"
        )


# ============================================================
# EXCEL
# ============================================================

def salvar_excel():

    if len(dados_excel) == 0:

        messagebox.showinfo(
            "Excel",
            "Ainda não existem dados efetivos para salvar."
        )

        return

    try:

        os.makedirs(
            PASTA_DADOS,
            exist_ok=True
        )

        agora = datetime.now()

        nome = (
            "ensaio_RC_"
            + agora.strftime(
                "%Y-%m-%d_%H-%M-%S"
            )
            + ".xlsx"
        )

        caminho = os.path.join(
            PASTA_DADOS,
            nome
        )

        workbook = Workbook()

        planilha = workbook.active

        planilha.title = "Dados"

        cabecalho = [
            "Ciclo",
            "Estado",
            "Tempo (s)",
            "Tensão (V)"
        ]

        planilha.append(
            cabecalho
        )

        for celula in planilha[1]:

            celula.font = Font(
                bold=True
            )

            celula.alignment = Alignment(
                horizontal="center"
            )

        for dado in dados_excel:

            planilha.append(
                [
                    dado["ciclo"],
                    dado["estado"],
                    dado["tempo"],
                    dado["tensao"]
                ]
            )

        # Largura das colunas

        larguras = [
            12,
            18,
            15,
            15
        ]

        for i, largura in enumerate(
            larguras,
            start=1
        ):

            planilha.column_dimensions[
                get_column_letter(i)
            ].width = largura

        # Congela cabeçalho

        planilha.freeze_panes = "A2"

        workbook.save(
            caminho
        )

        label_status.config(
            text=f"EXCEL SALVO: {nome}"
        )

        return caminho

    except Exception as erro:

        messagebox.showerror(
            "Erro ao salvar Excel",
            str(erro)
        )


# ============================================================
# CONEXÃO
# ============================================================

def conectar_arduino():

    global arduino

    if arduino is not None:

        try:

            if arduino.is_open:

                return True

        except:

            pass

    try:

        arduino = serial.Serial(
            PORTA,
            BAUDRATE,
            timeout=0.1
        )

        time.sleep(2)

        label_status.config(
            text=(
                f"STATUS: ARDUINO CONECTADO - {PORTA}"
            )
        )

        return True

    except Exception as erro:

        label_status.config(
            text="STATUS: ARDUINO DESCONECTADO"
        )

        messagebox.showerror(
            "Erro de conexão",
            (
                "Não foi possível conectar ao Arduino.\n\n"
                f"Porta: {PORTA}\n\n"
                f"{erro}\n\n"
                "Verifique a porta COM e se o Monitor Serial "
                "está fechado."
            )
        )

        return False


# ============================================================
# LIMPAR GRÁFICO
# ============================================================

def limpar_graficos():

    ax_carga.clear()

    ax_descarga.clear()

    ax_carga.set_title(
        "CARGA - Aguardando dados"
    )

    ax_carga.set_xlabel(
        "Tempo (s)"
    )

    ax_carga.set_ylabel(
        "Tensão (V)"
    )

    ax_carga.grid(True)

    ax_descarga.set_title(
        "DESCARGA - Aguardando dados"
    )

    ax_descarga.set_xlabel(
        "Tempo (s)"
    )

    ax_descarga.set_ylabel(
        "Tensão (V)"
    )

    ax_descarga.grid(True)

    figura.tight_layout()

    canvas.draw_idle()


# ============================================================
# DESENHAR CICLO ATUAL
# ============================================================

def desenhar_graficos():

    ax_carga.clear()

    ax_descarga.clear()

    # ========================================================
    # CARGA
    # ========================================================

    if dados_carga:

        tempos = [
            d[0]
            for d in dados_carga
        ]

        tensoes = [
            d[1]
            for d in dados_carga
        ]

        ax_carga.plot(
            tempos,
            tensoes,
            label="Capturada"
        )

    # ========================================================
    # CURVA TEÓRICA DE CARGA
    # ========================================================

    if tau > 0:

        tmax = max(
            tempo_5tau,
            max(
                [d[0] for d in dados_carga],
                default=0
            )
        )

        tmax *= 1.05

        teorico_t = []
        teorico_v = []

        for i in range(200):

            t = (
                tmax * i / 199
            )

            v = (
                VCC *
                (
                    1 -
                    math.exp(
                        -t / tau
                    )
                )
            )

            teorico_t.append(t)
            teorico_v.append(v)

        ax_carga.plot(
            teorico_t,
            teorico_v,
            "--",
            label="Teórica"
        )

    ax_carga.axvline(
        tempo_5tau,
        linestyle=":",
        label="5τ"
    )

    ax_carga.set_title(
        f"CARGA - Ciclo {ciclo_atual}"
    )

    ax_carga.set_xlabel(
        "Tempo (s)"
    )

    ax_carga.set_ylabel(
        "Tensão (V)"
    )

    ax_carga.set_ylim(
        0,
        5.2
    )

    ax_carga.grid(True)

    ax_carga.legend()


    # ========================================================
    # DESCARGA
    # ========================================================

    if dados_descarga:

        tempos = [
            d[0]
            for d in dados_descarga
        ]

        tensoes = [
            d[1]
            for d in dados_descarga
        ]

        ax_descarga.plot(
            tempos,
            tensoes,
            label="Capturada"
        )

    # ========================================================
    # CURVA TEÓRICA DE DESCARGA
    # ========================================================

    if tau > 0:

        tmax = max(
            tempo_5tau,
            max(
                [
                    d[0]
                    for d in dados_descarga
                ],
                default=0
            )
        )

        tmax *= 1.05

        teorico_t = []
        teorico_v = []

        for i in range(200):

            t = (
                tmax * i / 199
            )

            v = (
                VCC *
                math.exp(
                    -t / tau
                )
            )

            teorico_t.append(t)
            teorico_v.append(v)

        ax_descarga.plot(
            teorico_t,
            teorico_v,
            "--",
            label="Teórica"
        )

    ax_descarga.axvline(
        tempo_5tau,
        linestyle=":",
        label="5τ"
    )

    ax_descarga.set_title(
        f"DESCARGA - Ciclo {ciclo_atual}"
    )

    ax_descarga.set_xlabel(
        "Tempo (s)"
    )

    ax_descarga.set_ylabel(
        "Tensão (V)"
    )

    ax_descarga.set_ylim(
        0,
        5.2
    )

    ax_descarga.grid(True)

    ax_descarga.legend()

    figura.tight_layout()

    canvas.draw_idle()


# ============================================================
# PROCESSAR DADOS
# ============================================================

def processar_dado(
    ciclo,
    estado,
    tempo,
    tensao
):

    global ciclo_atual
    global estado_atual
    global dados_carga
    global dados_descarga
    global ultimo_ciclo_grafico

    ciclo_atual = ciclo
    estado_atual = estado

    # ========================================================
    # NOVO CICLO
    # ========================================================

    if ciclo != ultimo_ciclo_grafico:

        dados_carga.clear()

        dados_descarga.clear()

        ultimo_ciclo_grafico = ciclo

    # ========================================================
    # DADOS DO GRÁFICO
    # ========================================================

    if estado == "CARGA":

        dados_carga.append(
            (
                tempo,
                tensao
            )
        )

    elif estado == "DESCARGA":

        dados_descarga.append(
            (
                tempo,
                tensao
            )
        )

    # ========================================================
    # DADOS PARA EXCEL
    # ========================================================

    dados_excel.append(
        {
            "ciclo": ciclo,
            "estado": estado,
            "tempo": tempo,
            "tensao": tensao
        }
    )

    # ========================================================
    # TABELA
    # ========================================================

    janela.after(
        0,
        atualizar_interface,
        ciclo,
        estado,
        tempo,
        tensao
    )


# ============================================================
# ATUALIZAR INTERFACE
# ============================================================

def atualizar_interface(
    ciclo,
    estado,
    tempo,
    tensao
):

    tabela.insert(
        "",
        "end",
        values=(
            ciclo,
            estado,
            f"{tempo:.3f}",
            f"{tensao:.3f}"
        )
    )

    itens = tabela.get_children()

    if len(itens) > MAX_TABELA:

        tabela.delete(
            itens[0]
        )

    # ========================================================
    # Atualiza gráfico
    # ========================================================

    desenhar_graficos()

    # ========================================================
    # Auto-scroll
    # ========================================================

    ir_para_final()


# ============================================================
# LEITURA SERIAL
# ============================================================

def ler_serial():

    global teste_ativo

    while teste_ativo:

        try:

            if arduino is None:
                break

            linha = (
                arduino.readline()
                .decode(
                    "utf-8",
                    errors="ignore"
                )
                .strip()
            )

            if not linha:
                continue

            # =================================================
            # IGNORA MENSAGENS
            # =================================================

            if linha.startswith(
                "PRE,"
            ):
                continue

            if linha.startswith(
                "#"
            ):
                continue

            if linha in (
                "PRE_CONDICIONAMENTO",
                "PRE_CONDICIONAMENTO_OK",
                "ARDUINO_RC_PRONTO"
            ):
                continue

            # =================================================
            # DADOS
            # =================================================

            partes = linha.split(",")

            if len(partes) != 4:
                continue

            try:

                ciclo = int(
                    partes[0]
                )

                estado = (
                    partes[1]
                    .strip()
                    .upper()
                )

                tempo_ms = float(
                    partes[2]
                )

                tensao = float(
                    partes[3]
                )

            except:

                continue

            if estado not in (
                "CARGA",
                "DESCARGA"
            ):
                continue

            tempo = (
                tempo_ms / 1000.0
            )

            processar_dado(
                ciclo,
                estado,
                tempo,
                tensao
            )

        except Exception:

            time.sleep(
                0.05
            )


# ============================================================
# PARAR TESTE
# ============================================================

def parar_teste():

    global teste_ativo

    teste_ativo = False

    try:

        if (
            arduino is not None
            and arduino.is_open
        ):

            arduino.write(
                b"STOP\n"
            )

    except:

        pass

    label_status.config(
        text="STATUS: TESTE PARADO"
    )

    # Salva automaticamente

    if dados_excel:

        salvar_excel()


# ============================================================
# NOVO TESTE
# ============================================================

def novo_teste():

    global teste_ativo
    global ciclo_atual
    global estado_atual
    global ultimo_ciclo_grafico
    global dados_excel

    teste_ativo = False

    try:

        if (
            arduino is not None
            and arduino.is_open
        ):

            arduino.write(
                b"RESET\n"
            )

    except:

        pass

    dados_carga.clear()

    dados_descarga.clear()

    dados_excel.clear()

    ciclo_atual = 0

    estado_atual = "AGUARDANDO"

    ultimo_ciclo_grafico = 0

    for item in tabela.get_children():

        tabela.delete(item)

    limpar_graficos()

    label_status.config(
        text="STATUS: NOVO TESTE"
    )


# ============================================================
# SALVAR AUTOMATICAMENTE
# ============================================================

def salvar_automaticamente():

    if dados_excel:

        salvar_excel()


# ============================================================
# FECHAR
# ============================================================

def fechar_programa():

    global teste_ativo

    teste_ativo = False

    try:

        if (
            arduino is not None
            and arduino.is_open
        ):

            arduino.write(
                b"STOP\n"
            )

            time.sleep(
                0.2
            )

            arduino.close()

    except:

        pass

    if dados_excel:

        salvar_excel()

    janela.destroy()


janela.protocol(
    "WM_DELETE_WINDOW",
    fechar_programa
)


# ============================================================
# ATUALIZA CÁLCULOS AUTOMATICAMENTE
# ============================================================

entrada_capacitor.bind(
    "<KeyRelease>",
    lambda event: atualizar_calculos()
)

entrada_resistor.bind(
    "<KeyRelease>",
    lambda event: atualizar_calculos()
)

entrada_quantidade.bind(
    "<KeyRelease>",
    lambda event: atualizar_calculos()
)


# ============================================================
# INICIALIZAÇÃO
# ============================================================

atualizar_calculos()

limpar_graficos()

conectar_arduino()

# ============================================================
# INICIA LEITURA
# ============================================================

teste_ativo = True

thread_leitura = threading.Thread(
    target=ler_serial,
    daemon=True
)

thread_leitura.start()


# ============================================================
# LOOP
# ============================================================

janela.mainloop()